"""
Excel export helpers — gera arquivos .xlsx com gráficos de linha a partir dos
dados das séries temporais, e empacota CSV + XLSX em um único arquivo ZIP para
download.

Funções públicas:
    build_serie_xlsx(rows, metric_id, metric_name, escala) -> bytes
    build_series_xlsx(series_data)                         -> bytes
    build_datapoints_xlsx(rows)                            -> bytes
    make_zip_response(csv_rows, xlsx_bytes, base_filename) -> StreamingResponse
"""

import csv
import io
import zipfile
from collections import defaultdict

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Estilo ────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # azul escuro
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_CHART_W = 24   # largura do gráfico em cm
_CHART_H = 14   # altura do gráfico em cm


def _style_header(ws, n_cols: int) -> None:
    """Aplica preenchimento e fonte ao cabeçalho (linha 1)."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _col_width(name: str) -> float:
    return max(len(name) + 4, 12)


# ── Pivô + Gráfico ────────────────────────────────────────────────────────────

def _build_pivot(
    rows: list[dict],
    date_col: str,
    value_col: str,
    group_col: str,
) -> tuple[list, list[str], dict]:
    """
    Retorna (dates_raw, groups, pivot) onde:
      dates_raw  — lista de datas na ordem original (objetos date ou strings)
      groups     — lista de grupos únicos (ex: ['consolidado', 'controladora'])
      pivot      — {group: {str(date): valor}}
    """
    dates_raw = list(dict.fromkeys(r[date_col] for r in rows))
    groups = list(dict.fromkeys(r[group_col] for r in rows))
    pivot: dict[str, dict[str, float]] = defaultdict(dict)
    for r in rows:
        pivot[r[group_col]][str(r[date_col])] = r[value_col]
    return dates_raw, groups, pivot


def _write_pivot_and_chart(
    ws,
    rows: list[dict],
    date_col: str,
    value_col: str,
    group_col: str,
    title: str,
    y_title: str,
) -> None:
    """
    Escreve tabela pivô (datas × grupos) na planilha e adiciona gráfico de
    linhas abaixo dos dados.
    """
    if not rows:
        return

    dates_raw, groups, pivot = _build_pivot(rows, date_col, value_col, group_col)
    n_dates = len(dates_raw)
    n_cols = len(groups) + 1  # coluna A = Data, demais = um grupo cada

    # ── Cabeçalho ──
    ws.cell(row=1, column=1, value="Data")
    for col, grp in enumerate(groups, start=2):
        ws.cell(row=1, column=col, value=str(grp))

    # ── Dados ──
    for row_idx, d_raw in enumerate(dates_raw, start=2):
        ws.cell(row=row_idx, column=1, value=d_raw)
        for col, grp in enumerate(groups, start=2):
            val = pivot[grp].get(str(d_raw))
            if val is not None:
                ws.cell(row=row_idx, column=col, value=val)

    _style_header(ws, n_cols)

    ws.column_dimensions["A"].width = 14
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ── Gráfico de linhas ──
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = y_title
    chart.x_axis.title = "Data"
    chart.width = _CHART_W
    chart.height = _CHART_H

    data_ref = Reference(ws, min_col=2, max_col=n_cols, min_row=1, max_row=n_dates + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_dates + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    # Posiciona o gráfico 2 linhas abaixo dos dados
    chart_anchor = f"A{n_dates + 4}"
    ws.add_chart(chart, chart_anchor)


# ── Aba de dados brutos ───────────────────────────────────────────────────────

def _write_raw_sheet(ws, rows: list[dict]) -> None:
    """Escreve todos os campos de `rows` em formato tabular na planilha."""
    if not rows:
        return
    headers = list(rows[0].keys())
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
        ws.column_dimensions[get_column_letter(col)].width = _col_width(h)
    _style_header(ws, len(headers))
    for row_idx, row in enumerate(rows, start=2):
        for col, key in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col, value=row.get(key))


# ── Construtores de XLSX ──────────────────────────────────────────────────────

def build_serie_xlsx(
    rows: list[dict],
    metric_id: str,
    metric_name: str | None,
    escala: str,
) -> bytes:
    """
    XLSX para /serie/{metric_id}.
    Aba "Dados": tabela bruta.
    Aba "Gráfico": pivô datas × entidade + gráfico de linha.
    rows deve conter as chaves: data_iso, valor, entidade, escala_monetaria, …
    """
    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Dados"
    _write_raw_sheet(ws_raw, rows)

    ws_chart = wb.create_sheet("Gráfico")
    _write_pivot_and_chart(
        ws_chart, rows,
        date_col="data_iso",
        value_col="valor",
        group_col="entidade",
        title=metric_name or metric_id,
        y_title=f"Valor ({escala})",
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_series_xlsx(series_data: list[dict]) -> bytes:
    """
    XLSX para /series (batch de múltiplas métricas).
    Uma aba por métrica, cada aba com pivô datas × entidade + gráfico de linha.

    series_data: lista de dicts com chaves:
        metric_id, metric_name, escala, rows (list[dict] com data_iso/valor/entidade)
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove a aba padrão vazia

    for serie in series_data:
        mid: str = serie["metric_id"]
        mname: str | None = serie["metric_name"]
        escala: str = serie["escala"]
        rows: list[dict] = serie["rows"]

        # Nomes de aba: máximo 31 caracteres no Excel
        sheet_name = mid if len(mid) <= 31 else mid[:28] + "…"
        ws = wb.create_sheet(sheet_name)

        _write_pivot_and_chart(
            ws, rows,
            date_col="data_iso",
            value_col="valor",
            group_col="entidade",
            title=mname or mid,
            y_title=f"Valor ({escala})",
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_datapoints_xlsx(rows: list[dict]) -> bytes:
    """
    XLSX para /datapoints/flat.
    Aba "Dados": tabela bruta completa.
    Uma aba por metric_id com pivô period_end × entity_scope + gráfico de linha.
    rows deve ser list[DatapointResponse.model_dump()].
    """
    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "Dados"
    _write_raw_sheet(ws_raw, rows)

    # Agrupa por metric_id para gerar uma aba de gráfico por métrica
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[row.get("metric_id", "")].append(row)

    for mid, metric_rows in grouped.items():
        sheet_name = (mid[:31] if mid and len(mid) <= 31 else (mid[:28] + "…" if mid else "Métrica"))
        ws = wb.create_sheet(sheet_name)

        # Remapeia campos do DatapointResponse para o padrão do pivô
        pivot_rows = [
            {
                "data_iso": r.get("period_end"),
                "valor": r.get("value"),
                "entidade": r.get("entity_scope", ""),
            }
            for r in metric_rows
        ]
        escala = metric_rows[0].get("escala_monetaria", "") if metric_rows else ""

        _write_pivot_and_chart(
            ws, pivot_rows,
            date_col="data_iso",
            value_col="valor",
            group_col="entidade",
            title=mid,
            y_title=f"Valor ({escala})",
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Resposta ZIP ──────────────────────────────────────────────────────────────

def make_zip_response(
    csv_rows: list[dict],
    xlsx_bytes: bytes,
    base_filename: str,
) -> StreamingResponse:
    """
    Empacota CSV e XLSX em um arquivo ZIP e retorna como StreamingResponse.
    O CSV usa BOM UTF-8 (utf-8-sig) para abertura direta no Excel sem problemas
    de codificação de caracteres acentuados.
    """
    csv_buf = io.StringIO()
    if csv_rows:
        writer = csv.DictWriter(csv_buf, fieldnames=list(csv_rows[0].keys()))
        writer.writeheader()
        writer.writerows(csv_rows)
    csv_bytes = csv_buf.getvalue().encode("utf-8-sig")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{base_filename}.csv", csv_bytes)
        zf.writestr(f"{base_filename}.xlsx", xlsx_bytes)
    zip_buf.seek(0)

    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{base_filename}.zip"'},
    )
